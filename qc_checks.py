import re
import os
import pandas as pd
import numpy as np
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Removed logging.basicConfig - it's now handled by app.py
DATE_FORMAT = "%Y-%m-%d"

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


# ----------------------------- Helpers -----------------------------
def _find_column(df, candidates):
    """
    Case-insensitive lookup for a column in df.columns.
    candidates: list of possible header names (strings) from config.
    Returns first matching actual column name or None.
    """
    if not isinstance(candidates, list):
        candidates = [candidates] # Handle single-string entries
        
    lower_map = {c.lower().strip(): c for c in df.columns}
    for cand in candidates:
        if cand is None:
            continue
        key = cand.lower().strip()
        if key in lower_map:
            return lower_map[key]
    return None


def _is_present(val):
    """
    Treat numeric values (including 0) as present.
    For strings: strip whitespace and consider 'nan'/'none' as absent.
    None/NaN -> False.
    """
    if val is None:
        return False
    try:
        if pd.isna(val):
            return False
    except Exception:
        pass
    # Numeric -> present (including 0)
    if isinstance(val, (int, float)) and not (isinstance(val, float) and pd.isna(val)):
        return True
    s = str(val).strip()
    if s == "":
        return False
    if s.lower() in ("nan", "none","null", "n/a", "-","NAN","NONE","NULL","N/A"):
        return False
    return True


# ----------------------------- 1️ Detect Monitoring Period -----------------------------
def detect_period_from_rosco(rosco_path):
    """
    Attempts to find 'Monitoring Period' row anywhere in the Rosco file and extract two dates (YYYY-MM-DD).
    Returns (start_date, end_date) as pandas.Timestamp.
    Raises ValueError if not found or parsed.
    """
    # This function is heuristic-based and doesn't need config
    x = pd.read_excel(rosco_path, header=None, dtype=str)
    combined_text = x.fillna("").astype(str).apply(lambda row: " ".join(row.values), axis=1)
    match_rows = combined_text[combined_text.str.contains("Monitoring Period", case=False, na=False)]
    if match_rows.empty:
        match_rows = combined_text[combined_text.str.contains("Monitoring Periods|Monitoring period", case=False, na=False)]
    if match_rows.empty:
        all_text = " ".join(combined_text.tolist())
        found = re.findall(r"\d{4}-\d{2}-\d{2}", all_text)
        if len(found) >= 2:
            start_date = pd.to_datetime(found[0], format=DATE_FORMAT)
            end_date = pd.to_datetime(found[1], format=DATE_FORMAT)
            return start_date, end_date
        raise ValueError("Could not find 'Monitoring Period' text in Rosco file.")

    text_row = match_rows.iloc[0]
    found = re.findall(r"\d{4}-\d{2}-\d{2}", text_row)
    if len(found) >= 2:
        start_date = pd.to_datetime(found[0], format=DATE_FORMAT)
        end_date = pd.to_datetime(found[1], format=DATE_FORMAT)
        return start_date, end_date

    found_alt = re.findall(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", text_row)
    if len(found_alt) >= 2:
        try:
            start_date = pd.to_datetime(found_alt[0], dayfirst=False, errors="coerce")
            end_date = pd.to_datetime(found_alt[1], dayfirst=False, errors="coerce")
            if pd.notna(start_date) and pd.notna(end_date):
                return start_date, end_date
        except Exception:
            pass

    raise ValueError("Could not parse monitoring period dates from Rosco file.")


# ----------------------------- 2️ Load BSR -----------------------------
def detect_header_row(bsr_path, bsr_cols):
    df_sample = pd.read_excel(bsr_path, header=None, nrows=200)
    
    # Use config columns to find the header
    key_cols = [
        bsr_cols.get('market', ['market'])[0],
        bsr_cols.get('tv_channel', ['channel'])[0],
        bsr_cols.get('date', ['date'])[0],
        bsr_cols.get('start_time', ['start'])[0]
    ]
    
    for i, row in df_sample.iterrows():
        row_str = " ".join(row.dropna().astype(str).tolist()).lower()
        # Find row that contains several key column names
        if sum(col.lower() in row_str for col in key_cols) >= 2:
            return i
            
    raise ValueError("Could not detect header row in BSR file.")


def load_bsr(bsr_path, bsr_cols):
    header_row = detect_header_row(bsr_path, bsr_cols)
    df = pd.read_excel(bsr_path, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ----------------------------- 3️ Period Check -----------------------------
def period_check(df, start_date, end_date, bsr_cols):
    
    date_col = _find_column(df, bsr_cols.get('date', ['date']))
    
    if not date_col:
        logging.warning("Period Check: 'date' column not found.")
        df["Within_Period_OK"] = False
        df["Within_Period_Remark"] = "Date column not found"
        return df
        
    df["Date_checked"] = pd.to_datetime(df[date_col], errors="coerce").dt.date
    df["Within_Period_OK"] = df["Date_checked"].between(start_date.date(), end_date.date())
    df["Within_Period_Remark"] = df["Within_Period_OK"].apply(lambda x: "" if x else "Date outside monitoring period")
    df = df.drop(columns=["Date_checked"], errors="ignore")
    return df


# ----------------------------- 4️ Completeness Check -----------------------------
def completeness_check(df, bsr_cols, rules):
    
    # --- Map logical names to actual columns (from config) ---
    colmap = {
        "tv_channel": _find_column(df, bsr_cols['tv_channel']),
        "channel_id": _find_column(df, bsr_cols['channel_id']),
        "type_of_program": _find_column(df, bsr_cols['type_of_program']),
        "match_day": _find_column(df, bsr_cols['match_day']),
        "home_team": _find_column(df, bsr_cols['home_team']),
        "away_team": _find_column(df, bsr_cols['away_team']),
        "aud_estimates": _find_column(df, bsr_cols['aud_estimates']),
        "aud_metered": _find_column(df, bsr_cols['aud_metered']),
        "source": _find_column(df, bsr_cols['source'])
    }

    # --- Initialize result columns
    df["Completeness_OK"] = True
    df["Completeness_Remark"] = ""

    # --- Get rules from config ---
    live_types = set(rules.get('live_types', ['live', 'repeat', 'delayed']))
    relaxed_types = set(rules.get('relaxed_types', ['highlights']))

    # --- Iterate rows
    for idx, row in df.iterrows():
        missing = []

        # 1️ Mandatory Fields
        for logical, display in [
            ("tv_channel", "TV Channel"), 
            ("channel_id", "Channel ID"),
            ("match_day", "Match Day"), 
            ("source", "Source"),
            ("type_of_program", "Type of Program") # <-- THIS IS THE NEW LINE
        ]:
            colname = colmap.get(logical)
            if colname is None:
                missing.append(f"{display} (column not found)")
            elif not _is_present(row.get(colname)):
                missing.append(display)

        # 2️ Audience Logic
        aud_est_col = colmap.get("aud_estimates")
        aud_met_col = colmap.get("aud_metered")

        if not aud_est_col and not aud_met_col:
            missing.append("Audience (Estimates/Metered) (columns not found)")
        else:
            est_present = _is_present(row.get(aud_est_col)) if aud_est_col else False
            met_present = _is_present(row.get(aud_met_col)) if aud_met_col else False

            if not est_present and not met_present:
                missing.append("Both Audience fields are empty")
            elif est_present and met_present:
                missing.append("Both Audience fields are filled")

        # 3️ Type-based (Home/Away)
        type_col = colmap.get("type_of_program")
        prog_type = str(row.get(type_col) or "").strip().lower() if type_col else ""
        home_col, away_col = colmap.get("home_team"), colmap.get("away_team")

        if prog_type in live_types:
            if not home_col: missing.append("Home Team (column not found)")
            elif not _is_present(row.get(home_col)): missing.append("Home Team")
            
            if not away_col: missing.append("Away Team (column not found)")
            elif not _is_present(row.get(away_col)): missing.append("Away Team")

        elif prog_type not in relaxed_types:
            # Check for other types that *should* have teams
            # This logic is OK, but it only runs if prog_type is not empty
            if prog_type and not (prog_type in live_types or prog_type in relaxed_types):
                if home_col and not _is_present(row.get(home_col)): missing.append("Home Team")
                if away_col and not _is_present(row.get(away_col)): missing.append("Away Team")

        # 4️ Final result
        if missing:
            df.at[idx, "Completeness_OK"] = False
            df.at[idx, "Completeness_Remark"] = "; ".join(missing)
        else:
            df.at[idx, "Completeness_Remark"] = "All key fields present"

    return df

# ----------------------------- 5 Program Category Check -----------------------------

def parse_duration_to_minutes(duration_series):
    """Convert durations in HH:MM:SS or numeric to minutes"""
    results = []
    for item in duration_series:
        if pd.isna(item):
            results.append(np.nan)
            continue
        if isinstance(item, (int, float)):
            results.append(float(item))
            continue
        s = str(item).strip()
        try:
            num = float(s)
            results.append(num)
            continue
        except ValueError:
            pass
        parts = s.split(':')
        if len(parts) >= 2:
            try:
                hours = float(re.sub(r"[^0-9.]", "", parts[0]))
                minutes = float(re.sub(r"[^0-9.]", "", parts[1]))
                seconds = float(re.sub(r"[^0-9.]", "", parts[2])) if len(parts) >= 3 else 0.0
                total_minutes = (hours * 60) + minutes + (seconds / 60)
                results.append(total_minutes)
            except Exception:
                results.append(np.nan)
        else:
            results.append(np.nan)
    return pd.Series(results, index=duration_series.index)


def program_category_check(bsr_path, df, col_map, rules, file_rules):
    """Performs program category validation using fixture sheet"""
    # --- 1. Load Fixture ---
    try:
        xl = pd.ExcelFile(bsr_path)
        fixture_keyword = file_rules.get('fixture_sheet_keyword', 'fixture')
        fixture_sheet = next((s for s in xl.sheet_names if fixture_keyword in s.lower()), None)
        if not fixture_sheet:
            df["Program_Category_OK"] = False
            df["Program_Category_Remark"] = "Fixture list sheet missing"
            return df
        df_fix = xl.parse(fixture_sheet)
    except Exception as e:
        df["Program_Category_OK"] = False
        df["Program_Category_Remark"] = f"Error loading fixture sheet: {e}"
        return df

    # --- 2. Column Resolution ---
    def _find_column(dframe, options):
        if not options: return None
        if isinstance(options, list):
            for o in options:
                if o in dframe.columns: return o
        elif options in dframe.columns: return options
        return None

    bsr_cols = col_map['bsr']
    fix_cols = col_map['fixture']

    col_home_bsr = _find_column(df, bsr_cols['home_team'])
    col_away_bsr = _find_column(df, bsr_cols['away_team'])
    col_date_bsr = _find_column(df, bsr_cols['date'])
    col_progtype = _find_column(df, bsr_cols['type_of_program'])
    col_desc = _find_column(df, bsr_cols['program_desc'])
    col_source = _find_column(df, bsr_cols['source'])
    col_start_utc = _find_column(df, bsr_cols['start_time'])
    col_end_utc = _find_column(df, bsr_cols['end_time'])
    col_duration_direct = _find_column(df, bsr_cols['duration'])

    col_home_fix = _find_column(df_fix, fix_cols['home_team'])
    col_away_fix = _find_column(df_fix, fix_cols['away_team'])
    col_date_fix = _find_column(df_fix, fix_cols['date'])
    col_start_fix = _find_column(df_fix, fix_cols['start_time'])

    # --- 3. Prepare Dates/Times ---
    df.columns = df.columns.map(str)
    df_fix.columns = df_fix.columns.map(str)
    base_date_str = df[col_date_bsr].astype(str)

    for c in [col_start_utc, col_end_utc]:
        if c:
            direct_dt = pd.to_datetime(df[c], errors='coerce')
            combined_dt = pd.to_datetime(base_date_str + ' ' + df[c].astype(str), errors='coerce')
            df[f"_dt_{c}"] = direct_dt.combine_first(combined_dt)

    # Fix fixture times
    if col_date_fix and col_start_fix:
        df_fix[col_date_fix] = pd.to_datetime(df_fix[col_date_fix], errors='coerce')
        df_fix[col_start_fix] = pd.to_datetime(
            df_fix[col_date_fix].dt.strftime('%Y-%m-%d') + ' ' + df_fix[col_start_fix].astype(str),
            errors='coerce'
        )

    # ----------------------- 4. Duration Calculation --------------------
    duration_calc = (df[f"_dt_{col_end_utc}"] - df[f"_dt_{col_start_utc}"]).dt.total_seconds() / 60
    duration_direct = parse_duration_to_minutes(df[col_duration_direct]) if col_duration_direct else pd.Series(np.nan, index=df.index)
    df['duration_min'] = duration_calc.combine_first(duration_direct)
    df['_bsr_start_time'] = df.get(f"_dt_{col_start_utc}", pd.NaT)

    # --- 5. Clean Team Names ---
    df['home_clean'] = df[col_home_bsr].astype(str).str.strip().str.lower()
    df['away_clean'] = df[col_away_bsr].astype(str).str.strip().str.lower()
    df_fix['home_clean'] = df_fix[col_home_fix].astype(str).str.strip().str.lower()
    df_fix['away_clean'] = df_fix[col_away_fix].astype(str).str.strip().str.lower()

    # ---- 6. Initialize Columns ---
    df["Program_Category_Expected"] = pd.NA
    df["Program_Category_Actual"] = df[col_progtype].astype(str).str.strip().str.lower()
    df["Program_Category_OK"] = False
    df["Program_Category_Remark"] = pd.NA

    # --- 7. Load Rule Parameters ---
    live_tolerance = rules.get('live_tolerance_min', 30)
    bsa_max_duration = rules.get('bsa_max_duration', 180)
    support_min = rules.get('support_duration_min', 10)
    support_max = rules.get('support_duration_max', 60)
    match_types = set(rules.get('live_types', []))
    magazine_types = set(rules.get('relaxed_types', []))

    # --- 8. Fixture Matching Logic ---
    unprocessed_indices = df.index[df["Program_Category_Actual"].isin(match_types)]

    for _, fix_row in df_fix.iterrows():
        home_fix, away_fix, start_fix, date_fix = fix_row.get('home_clean'), fix_row.get('away_clean'), fix_row.get(col_start_fix), fix_row.get(col_date_fix)
        if pd.isna(start_fix) or not home_fix or not away_fix:
            continue

        # Match BSR entries to fixture
        matches = df.loc[
            (df.index.isin(unprocessed_indices)) &
            (df['home_clean'] == home_fix) &
            (df['away_clean'] == away_fix)
        ]

        if matches.empty:
            continue

        matches = matches[pd.to_datetime(matches[col_date_bsr], errors='coerce').dt.date == pd.to_datetime(date_fix).date()]
        if matches.empty:
            continue

        matches = matches.sort_values(by=['_bsr_start_time'],kind="mergesort")
        first_idx = matches.index[0]
        first_start = matches.iloc[0]['_bsr_start_time']
        first_dur = matches.iloc[0]['duration_min']

        if pd.isna(first_start):
            df.loc[matches.index, 'Program_Category_Expected'] = 'unknown'
            df.loc[matches.index, 'Program_Category_Remark'] = 'Invalid BSR start time'
            continue

        start_diff = (first_start - start_fix).total_seconds() / 60
        live_min, live_max = bsa_max_duration - live_tolerance, bsa_max_duration + live_tolerance

        if (abs(start_diff) <= live_tolerance) and (pd.notna(first_dur) and live_min <= first_dur <= live_max):
            df.loc[first_idx, 'Program_Category_Expected'] = 'live'
        else:
            df.loc[first_idx, 'Program_Category_Expected'] = 'delayed'

        if len(matches) > 1:
            df.loc[matches.index[1:], 'Program_Category_Expected'] = 'repeat'

        unprocessed_indices = unprocessed_indices.difference(matches.index)

    # --- 9. Verification & Remarks ---
    for idx, row in df.iterrows():
        actual, expected, duration = row["Program_Category_Actual"], row["Program_Category_Expected"], row["duration_min"]
        ok, remark = False, ""

        if actual in magazine_types:
            df.at[idx, "Program_Category_Expected"] = actual
            if pd.isna(duration):
                remark = "Invalid duration"
            elif support_min <= duration <= support_max:
                ok, remark = True, "OK"
            else:
                remark = f"Invalid duration ({duration:.2f} min)"
        elif actual in match_types:
            if pd.isna(expected):
                remark = "No matching fixture found"
                df.at[idx, "Program_Category_Expected"] = "unknown"
            elif actual == expected:
                ok, remark = True, "OK"
            else:
                remark = f"Expected '{expected}', found '{actual}'"
        else:
            remark = f"Invalid Actual Type: {actual}"

        df.at[idx, "Program_Category_OK"] = ok
        df.at[idx, "Program_Category_Remark"] = remark

    # --- 10. Cleanup ---
    df = df.drop(columns=[c for c in [
        'duration_min', '_bsr_start_time', 'home_clean', 'away_clean',
        f"_dt_{col_start_utc}", f"_dt_{col_end_utc}"
    ] if c in df.columns], errors='ignore')

    return df


# ----------------------------- 6 Event / Matchday / Competition Check -----------------------------
def check_event_matchday_competition(df, bsr_path, col_map, file_rules):

    logging.info("Starting Event / Matchday / Fixture consistency check...")
    
    bsr_cols = col_map['bsr']
    fix_cols = col_map['fixture']
    
    col_progtype = _find_column(df, bsr_cols['type_of_program'])
    if not col_progtype:
        logging.error(" 'Type of program' column not found. Skipping Event check.")
        df["Event_Matchday_OK"] = False
        df["Event_Matchday_Remark"] = "Error: 'Type of program' column not found."
        return df

    # Default for all rows - will be overwritten for matches
    df["Event_Matchday_OK"] = pd.NA 
    df["Event_Matchday_Remark"] = "Not applicable for this program type"

    # --- Load fixture list ---
    fixture_df = None
    try:
        excel_file = pd.ExcelFile(bsr_path)
        fixture_keyword = file_rules.get('fixture_sheet_keyword', 'fixture')
        fixture_sheet = next((s for s in excel_file.sheet_names if fixture_keyword in s.lower()), None)
        
        if fixture_sheet:
            fixture_df = excel_file.parse(fixture_sheet)
            fixture_df.columns = [c.strip() for c in fixture_df.columns]
        else:
            logging.warning(" No sheet containing 'fixture' found.")
    except Exception as e:
        logging.error(f" Error loading fixture list: {e}")

    # --- Normalize fixture data ---
    if fixture_df is not None:
        fix_event_col = _find_column(fixture_df, fix_cols['event'])
        fix_home_col = _find_column(fixture_df, fix_cols['home_team'])
        fix_away_col = _find_column(fixture_df, fix_cols['away_team'])
        fix_md_col = _find_column(fixture_df, fix_cols['match_day'])

        for col in [fix_event_col, fix_home_col, fix_away_col, fix_md_col]:
            if col:
                fixture_df[col] = fixture_df[col].astype(str).str.strip().str.lower()
            else:
                logging.warning(f" Fixture list missing a key column. Check config.")
                fixture_df = None 
                break

    # --- Find BSR columns ---
    bsr_event_col = _find_column(df, bsr_cols['event'])
    bsr_home_col = _find_column(df, bsr_cols['home_team'])
    bsr_away_col = _find_column(df, bsr_cols['away_team'])
    bsr_md_col = _find_column(df, bsr_cols['match_day'])

    # --- Define all program types that are full matches ---
    # This matches your flowchart's first check
    match_program_types = ['live', 'repeat', 'delayed']

    for i, row in df.iterrows():
        try:
            prog_type = str(row.get(col_progtype, "")).strip().lower()

            # --- This is the "YES" branch of your flowchart ---
            if prog_type in match_program_types:
                if fixture_df is None:
                    df.at[i, "Event_Matchday_OK"] = False
                    df.at[i, "Event_Matchday_Remark"] = "Fixture list missing or invalid"
                    continue

                event = str(row.get(bsr_event_col, "")).strip().lower()
                home = str(row.get(bsr_home_col, "")).strip().lower()
                away = str(row.get(bsr_away_col, "")).strip().lower()
                matchday = str(row.get(bsr_md_col, "")).strip().lower()

                if not event or not home or not away or not matchday:
                    df.at[i, "Event_Matchday_OK"] = False
                    df.at[i, "Event_Matchday_Remark"] = "Missing event/home/away/matchday in BSR"
                    continue

                # Check if this match exists in the fixture list
                match = fixture_df[
                    (fixture_df[fix_event_col] == event)
                    & (fixture_df[fix_home_col] == home)
                    & (fixture_df[fix_away_col] == away)
                    & (fixture_df[fix_md_col] == matchday)
                ]

                # This is the "Match Found?" check
                if match.empty:
                    df.at[i, "Event_Matchday_OK"] = False
                    df.at[i, "Event_Matchday_Remark"] = "No matching fixture found"
                else:
                    df.at[i, "Event_Matchday_OK"] = True
                    df.at[i, "Event_MatchMday_Remark"] = "Fixture found"
            
            # --- This is the "NO" branch of your flowchart ---
            # (If prog_type is not a match, it keeps the default "Not applicable" values)

        except Exception as e:
            df.at[i, "Event_Matchday_OK"] = False
            df.at[i, "Event_Matchday_Remark"] = f"Error: {e}"

    logging.info("✅ Event / Matchday / Fixture consistency check completed.")
    return df

# ------------------------7 Market channel consistency check-----------------------------------
def market_channel_consistency_check(df_bsr, rosco_path, col_map, file_rules):
    
    logging.info(" Starting Market & Channel Consistency Check...")
    
    bsr_cols = col_map['bsr']
    rosco_cols = col_map['rosco']
    
    # --- Normalization helper for ROSCO ---
    def normalize_channel(name):
        if pd.isna(name): return ""
        s = str(name)
        s = re.sub(r"\(.*?\)|\[.*?\]", "", s)
        s = re.split(r"[-–—]", s)[0]
        s = re.sub(r"[^0-9a-zA-Z\s]", " ", s)
        s = re.sub(r"\s+", " ", s).strip().lower()
        return s

    # --- Load ROSCO reference sheet ---
    rosco_df = None
    if rosco_path:
        try:
            xls = pd.ExcelFile(rosco_path)
            ignore_sheet = file_rules.get('rosco_ignore_sheet', 'general')
            sheet_name = next((s for s in xls.sheet_names if ignore_sheet not in s.lower()), None)
            if sheet_name:
                rosco_df = xls.parse(sheet_name)
            else:
                logging.warning(f" No valid sheet found in ROSCO (ignoring '{ignore_sheet}').")
        except Exception as e:
            logging.error(f" Error loading ROSCO file: {e}")
            df_bsr["Market_Channel_Consistency_OK"] = False
            df_bsr["Market_Channel_Program_Remark"] = f"Error loading ROSCO: {e}"
            return df_bsr

    # --- Build valid (Market, Channel) pairs from ROSCO ---
    valid_pairs = set()
    rosco_country_col = rosco_cols.get('channel_country', 'ChannelCountry')
    rosco_name_col = rosco_cols.get('channel_name', 'ChannelName')
    
    if rosco_df is not None:
        if {rosco_country_col, rosco_name_col}.issubset(rosco_df.columns):
            for _, row in rosco_df.iterrows():
                market = str(row[rosco_country_col]).strip().lower()
                channel = normalize_channel(row[rosco_name_col])
                if market and channel:
                    valid_pairs.add((market, channel))
            logging.info(f" Loaded {len(valid_pairs)} valid Market+Channel pairs from ROSCO.")
        else:
            logging.warning(f" '{rosco_country_col}' or '{rosco_name_col}' not in ROSCO sheet.")

    # --- Prepare result columns ---
    df_bsr["Market_Channel_Consistency_OK"] = True
    df_bsr["Market_Channel_Program_Remark"] = "OK"
    
    # --- Find BSR columns ---
    bsr_market_col = _find_column(df_bsr, bsr_cols['market'])
    bsr_channel_col = _find_column(df_bsr, bsr_cols['tv_channel'])
    
    if not bsr_market_col or not bsr_channel_col:
        logging.error(" Market/Channel Check: BSR columns not found. Skipping.")
        df_bsr["Market_Channel_Consistency_OK"] = False
        df_bsr["Market_Channel_Program_Remark"] = "BSR columns not found"
        return df_bsr

    # --- Validate each row in BSR ---
    for idx, row in df_bsr.iterrows():
        remarks = []
        market = str(row.get(bsr_market_col, "")).strip().lower()
        channel = str(row.get(bsr_channel_col, "")).strip()

        if not market or not channel:
            df_bsr.at[idx, "Market_Channel_Consistency_OK"] = False
            remarks.append("Missing market or channel")
        elif valid_pairs:
            if (market, normalize_channel(channel)) not in valid_pairs:
                df_bsr.at[idx, "Market_Channel_Consistency_OK"] = False
                remarks.append("Market+Channel not found in ROSCO")

        df_bsr.at[idx, "Market_Channel_Program_Remark"] = "; ".join(remarks) if remarks else "OK"

    logging.info(" Market & Channel Consistency Check completed.")
    return df_bsr

# --------------------------8 Domestic market check---------------------------------
def domestic_market_check(df, project_config, bsr_cols, debug=False):
    
    league_name = project_config.get('league_keyword', 'F24 Spain')
    domestic_market = project_config.get('domestic_market', 'Spain')
    domestic_keywords = project_config.get('domestic_league_keywords', ['F24 Spain'])
    
    logging.info(f" Running domestic market coverage check for league: {league_name}")

    # --- Find columns using config mapping ---
    market_col = _find_column(df, bsr_cols['market'])
    competition_col = _find_column(df, bsr_cols['competition'])
    event_col = _find_column(df, bsr_cols['event'])
    program_type_col = _find_column(df, bsr_cols['type_of_program'])
    matchday_col = _find_column(df, bsr_cols['match_day'])

    required_cols_found = [market_col, competition_col, event_col, program_type_col, matchday_col]
    
    if any(col is None for col in required_cols_found):
        logging.warning("Domestic Market Check: Missing required column. Skipping.")
        df["Domestic_Market_Coverage_Check_OK"] = pd.NA
        df["Domestic Market Coverage Remark"] = "Required column missing"
        return df

    # Normalize text (on found columns)
    for col in required_cols_found:
        df[col] = df[col].astype(str).str.strip()

    # --- Use config variables instead of hard-coded strings ---
    df["is_domestic_market"] = df[market_col].str.contains(domestic_market, case=False, na=False)
    
    df["is_target_league"] = df[competition_col].apply(
        lambda x: any(kw.lower() in str(x).lower() for kw in domestic_keywords)
    ) | df[event_col].apply(
        lambda x: any(kw.lower() in str(x).lower() for kw in domestic_keywords)
    )

    # Initialize output columns
    df["Domestic_Market_Coverage_Check_OK"] = pd.NA
    df["Domestic_Market_Coverage_Check_OK"] = df["Domestic_Market_Coverage_Check_OK"].astype('object')
    df["Domestic Market Coverage Remark"] = "Not Applicable"

    target_rows = df[df["is_target_league"] & df["is_domestic_market"]]
    if target_rows.empty:
        logging.warning(f" No '{league_name}' entries found for '{domestic_market}' market.")
        df.drop(columns=["is_domestic_market", "is_target_league"], inplace=True, errors="ignore")
        return df

    all_matchdays = target_rows[matchday_col].unique()
    if debug:
        logging.info(f" Found {len(all_matchdays)} matchdays for {domestic_market} market: {all_matchdays}")

    for md in all_matchdays:
        if not md or str(md).lower() == 'nan':
            continue
            
        md_rows = target_rows[target_rows[matchday_col] == md]
        live_present = any(md_rows[program_type_col].str.contains("Live", case=False, na=False))
        delayed_present = any(md_rows[program_type_col].str.contains("Delayed", case=False, na=False))

        condition = (df[matchday_col] == md) & df["is_target_league"] & df["is_domestic_market"]

        if not live_present and not delayed_present:
            df.loc[condition, "Domestic_Market_Coverage_Check_OK"] = False
            df.loc[condition, "Domestic Market Coverage Remark"] = f"No live/delayed coverage for matchday {md}"
        else:
            df.loc[condition, "Domestic_Market_Coverage_Check_OK"] = True
            remark = "Live" if live_present else ""
            remark += " & " if live_present and delayed_present else ""
            remark += "Delayed" if delayed_present else ""
            df.loc[condition, "Domestic Market Coverage Remark"] = f"{remark} coverage present for matchday {md}"

    # Set non-applicable rows
    mask_highlights = df[program_type_col].str.contains("Highlight|Magazine", case=False, na=False) & df["is_domestic_market"]
    df.loc[mask_highlights, "Domestic_Market_Coverage_Check_OK"] = pd.NA
    df.loc[mask_highlights, "Domestic Market Coverage Remark"] = "Not applicable for highlights or magazine programs"

    df.drop(columns=["is_domestic_market", "is_target_league"], inplace=True, errors="ignore")
    return df

# ----------------------------9 Rates and Ratings Check-------------------------------
def rates_and_ratings_check(df, bsr_cols):
    
    est_col = _find_column(df, bsr_cols['aud_estimates'])
    met_col = _find_column(df, bsr_cols['aud_metered'])
    
    if est_col is None:
        df[est_col] = pd.NA # Create dummy column to avoid errors
        logging.warning("Rates/Ratings Check: Audience Estimates column not found.")
    if met_col is None:
        df[met_col] = pd.NA
        logging.warning("Rates/Ratings Check: Audience Metered column not found.")

    present_est = df[est_col].apply(_is_present)
    present_met = df[met_col].apply(_is_present)

    both_empty_mask = (~present_est) & (~present_met)
    both_present_mask = (present_est) & (present_met)
    exactly_one_mask = (present_est ^ present_met)

    df["Rates_Ratings_QC_OK"] = True
    df["Rates_Ratings_QC_Remark"] = ""
    
    df.loc[both_empty_mask, "Rates_Ratings_QC_OK"] = False
    df.loc[both_empty_mask, "Rates_Ratings_QC_Remark"] = "Missing audience ratings (both empty)"
    
    df.loc[both_present_mask, "Rates_Ratings_QC_OK"] = False
    df.loc[both_present_mask, "Rates_Ratings_QC_Remark"] = "Invalid: both metered and estimated present"
    
    df.loc[exactly_one_mask, "Rates_Ratings_QC_OK"] = True
    df.loc[exactly_one_mask, "Rates_Ratings_QC_Remark"] = "Valid: one rating source available"

    return df

# -----------------------------10 DUPLICATED MARKET CHECK -----------------------------
def duplicated_market_check(df_bsr, macro_path, project, col_map, file_rules, debug=False):
    """
    Validates duplication of feeds across markets using Macro file rules.
    Keeps existing logic (origin/duplicate validation) but now also returns
    a list of duplicated Channel IDs used across multiple markets.
    """
    result_col = "Duplicated_Markets_Check_OK"
    remark_col = "Duplicated_Markets_Remark"

    df_bsr[result_col] = pd.NA
    df_bsr[result_col] = df_bsr[result_col].astype('object')
    df_bsr[remark_col] = "Not Applicable"

    league_keyword = project.get('league_keyword', 'F24 Spain')
    bsr_cols = col_map['bsr']
    macro_cols = col_map['macro']

    duplicated_channels = set()

    if not macro_path or not os.path.exists(macro_path):
        df_bsr[remark_col] = "Macro file missing"
        return df_bsr, list(duplicated_channels)

    try:
        # --- Load and clean Macro Data ---
        macro_sheet = file_rules.get('macro_sheet_name', 'Data Core')
        header_row = file_rules.get('macro_header_row', 1)
        macro_df = pd.read_excel(macro_path, sheet_name=macro_sheet, header=header_row, dtype=str)
        macro_df.columns = macro_df.columns.str.strip()

        # Find macro columns
        proj_col = macro_cols['projects']
        orig_mkt_col = macro_cols['orig_market']
        orig_ch_col = macro_cols['orig_channel']
        dup_mkt_col = macro_cols['dup_market']
        dup_ch_col = macro_cols['dup_channel']
        
        macro_df = macro_df[
            macro_df[proj_col].astype(str).str.contains(league_keyword, case=False, na=False)
        ].copy()

        if macro_df.empty:
            df_bsr[remark_col] = f"No duplication rules found for {league_keyword}"
            return df_bsr, list(duplicated_channels)

        for col in [orig_mkt_col, orig_ch_col, dup_mkt_col, dup_ch_col]:
            macro_df[col] = macro_df[col].astype(str).str.strip().str.lower()

        # --- Find BSR columns ---
        mkt_col = _find_column(df_bsr, bsr_cols['market'])
        ch_col = _find_column(df_bsr, bsr_cols['tv_channel'])
        comp_col = _find_column(df_bsr, bsr_cols['competition'])
        evt_col = _find_column(df_bsr, bsr_cols['event'])

        # --- Filter BSR for selected league (competition/event) ---
        in_league = (
            df_bsr[comp_col].astype(str).str.lower().str.contains(league_keyword.lower(), na=False)
            | df_bsr[evt_col].astype(str).str.lower().str.contains(league_keyword.lower(), na=False)
        )
        df_league = df_bsr[in_league].copy()

        if df_league.empty:
            df_bsr[remark_col] = f"No events found for {league_keyword}"
            return df_bsr, list(duplicated_channels)

        # --- Core Duplication Logic ---
        for _, row in macro_df.iterrows():
            orig_market = row[orig_mkt_col]
            orig_channel = row[orig_ch_col]
            dup_market = row[dup_mkt_col]
            dup_channel = row[dup_ch_col]

            # Track duplicated channels for overlap check
            duplicated_channels.add(orig_channel)
            duplicated_channels.add(dup_channel)

            orig_events = set(df_league[
                (df_league[mkt_col].astype(str).str.lower() == orig_market)
                & (df_league[ch_col].astype(str).str.lower() == orig_channel)
            ][evt_col])

            dup_events = set(df_league[
                (df_league[mkt_col].astype(str).str.lower() == dup_market)
                & (df_league[ch_col].astype(str).str.lower() == dup_channel)
            ][evt_col])

            if not orig_events:
                status = pd.NA
                remark = f"No events found in {orig_market} / {orig_channel}"
            elif orig_events.issubset(dup_events):
                status = True
                remark = f"All events correctly duplicated to {dup_market} / {dup_channel}"
            else:
                missing = orig_events - dup_events
                status = False
                remark = f"Missing {len(missing)} events in {dup_market} / {dup_channel}"

            # Apply results to all relevant rows
            orig_rows_mask = (df_bsr[mkt_col].astype(str).str.lower() == orig_market) & \
                             (df_bsr[ch_col].astype(str).str.lower() == orig_channel) & in_league
            dup_rows_mask = (df_bsr[mkt_col].astype(str).str.lower() == dup_market) & \
                            (df_bsr[ch_col].astype(str).str.lower() == dup_channel) & in_league

            df_bsr.loc[orig_rows_mask | dup_rows_mask, result_col] = status
            df_bsr.loc[orig_rows_mask | dup_rows_mask, remark_col] = remark

        if debug:
            print(f" Duplicated Market Check completed. Found {len(duplicated_channels)} duplicated channels across markets.")

        return df_bsr, list(duplicated_channels)

    except Exception as e:
        df_bsr[result_col] = False
        df_bsr[remark_col] = str(e)
        return df_bsr, list(duplicated_channels)
    
# -----------------------------11 OVERLAP / DUPLICATE ROW / DAYBREAK CHECK -----------------------------
def overlap_duplicate_daybreak_check(df, bsr_cols, rules, duplicated_channels=None):
    """
    Combined check for:
      - Overlap (only within same Market+Channel; end==next_start is NOT overlap)
      - Duplicate rows (exact duplicates within same Market+Channel only)
      - Daybreak/continuation checks (sensible rules across midnight)
    """
    df_in = df.copy(deep=True)
    duplicated_channels = set(duplicated_channels or [])

    # --- find columns ---
    col_market = _find_column(df_in, bsr_cols.get('market'))
    col_channel = _find_column(df_in, bsr_cols.get('tv_channel'))
    col_channel_id = _find_column(df_in, bsr_cols.get('channel_id'))
    col_date = _find_column(df_in, bsr_cols.get('date'))
    col_start = _find_column(df_in, bsr_cols.get('start_time'))
    col_end = _find_column(df_in, bsr_cols.get('end_time'))
    col_pay = _find_column(df_in, bsr_cols.get('pay_tv'))
    col_event = _find_column(df_in, bsr_cols.get('event'))
    ignore_platforms = rules.get('ignore_platforms', [])

    # required columns
    if not col_channel_id and not col_channel:
        logging.error("Overlap check: missing channel identifier in BSR columns.")
    if not col_start or not col_end:
        df_in["Overlap_OK"] = False
        df_in["Overlap_Remark"] = "Check skipped: missing start or end columns"
        df_in["Duplicate_OK"] = False
        df_in["Duplicate_Remark"] = "Check skipped: missing start or end columns"
        df_in["Daybreak_OK"] = False
        df_in["Daybreak_Remark"] = "Check skipped: missing start or end columns"
        return df_in

    # --- build full datetimes using date + time ---
    def combine_date_time(date_series, time_series):
        combined = pd.Series([pd.NaT]*len(time_series), index=time_series.index)
        if date_series is not None and date_series.name in df_in.columns:
            try:
                date_part = pd.to_datetime(df_in[date_series.name], errors='coerce').dt.strftime('%Y-%m-%d')
                combined = pd.to_datetime(date_part + ' ' + df_in[time_series.name].astype(str), errors='coerce')
            except Exception:
                combined = pd.to_datetime(df_in[time_series.name], errors='coerce')
        else:
            combined = pd.to_datetime(df_in[time_series.name], errors='coerce')

        still_na = combined.isna()
        if still_na.any():
            combined.loc[still_na] = pd.to_datetime(df_in.loc[still_na, time_series.name], errors='coerce')
        return combined

    date_series = df_in[col_date] if col_date else None
    df_in['_start_dt'] = combine_date_time(date_series, df_in[col_start])
    df_in['_end_dt']   = combine_date_time(date_series, df_in[col_end])

    # initialize outputs
    overlap_ok = pd.Series(True, index=df_in.index)
    overlap_remark = pd.Series("OK", index=df_in.index)
    duplicate_ok = pd.Series(True, index=df_in.index)
    duplicate_remark = pd.Series("OK", index=df_in.index)
    daybreak_ok = pd.Series(True, index=df_in.index)
    daybreak_remark = pd.Series("OK", index=df_in.index)

    # ignored platforms
    is_ignored_platform = pd.Series(False, index=df_in.index)
    if col_pay and ignore_platforms:
        try:
            pattern = '|'.join([p.lower() for p in ignore_platforms])
            is_ignored_platform = df_in[col_pay].astype(str).str.lower().str.contains(pattern, na=False)
        except:
            is_ignored_platform = pd.Series(False, index=df_in.index)

    # sorting
    sort_by = []
    if col_market: sort_by.append(col_market)
    if col_channel_id: sort_by.append(col_channel_id)
    elif col_channel: sort_by.append(col_channel)
    sort_by.append('_start_dt')

    df_work = df_in.sort_values(by=sort_by, na_position='last').reset_index()

    tol = rules.get('daybreak_gap_tolerance_min', 2)

    # === Overlap / Daybreak ===
    group_cols = []
    if col_market: group_cols.append(col_market)
    if col_channel_id: group_cols.append(col_channel_id)
    else: group_cols.append(col_channel)

    for _, grp in df_work.groupby(group_cols, sort=False):
        grp = grp.sort_values('_start_dt', na_position='last')
        prev_end = None

        # determine if this group should be skipped
        chan_val = grp.iloc[0][col_channel_id] if col_channel_id else grp.iloc[0][col_channel]
        skip_this_group = (str(chan_val).strip().lower() in {str(x).strip().lower() for x in duplicated_channels})

        for _, row in grp.iterrows():
            idx_orig = row['index']
            start = row['_start_dt']
            end = row['_end_dt']

            if pd.isna(start) or pd.isna(end):
                overlap_ok.at[idx_orig] = False
                overlap_remark.at[idx_orig] = "Invalid start or end time"
                prev_end = end if not pd.isna(end) else prev_end
                continue

            if skip_this_group:
                overlap_ok.at[idx_orig] = True
                overlap_remark.at[idx_orig] = "Skipped (channel duplicated across markets)"
                prev_end = end
                continue

            # actual overlap check
            if prev_end is not None and not pd.isna(prev_end):
                gap_min = (start - prev_end).total_seconds() / 60.0

                if gap_min < 0:
                    overlap_ok.at[idx_orig] = False
                    overlap_remark.at[idx_orig] = "Overlap detected with previous program"
                else:
                    overlap_ok.at[idx_orig] = True
                    overlap_remark.at[idx_orig] = "OK"

                # daybreak
                if gap_min > tol:
                    daybreak_ok.at[idx_orig] = False
                    daybreak_remark.at[idx_orig] = f"Large gap from previous program ({gap_min:.1f} min) > tolerance {tol} min"
                else:
                    daybreak_ok.at[idx_orig] = True
                    daybreak_remark.at[idx_orig] = "OK"
            else:
                overlap_ok.at[idx_orig] = True
                overlap_remark.at[idx_orig] = "OK"
                daybreak_ok.at[idx_orig] = True
                daybreak_remark.at[idx_orig] = "OK"

            prev_end = end

    # ============================
    # DUPLICATE DETECTION
    # ============================

    # we use NORMALIZED datetime columns instead of raw strings
    dup_subset = []
    if col_market: dup_subset.append(col_market)
    if col_channel_id: dup_subset.append(col_channel_id)
    else: dup_subset.append(col_channel)
    dup_subset.append("_start_dt")
    dup_subset.append("_end_dt")
    if col_event: dup_subset.append(col_event)

    try:
        raw_dups = df_in.duplicated(subset=dup_subset, keep=False)

        for idx in df_in[raw_dups].index:
            row = df_in.loc[idx]

            mask = (
                (df_in[col_market] == row[col_market]) &
                ((df_in[col_channel_id] == row[col_channel_id]) if col_channel_id
                 else (df_in[col_channel] == row[col_channel])) &
                (df_in["_start_dt"] == row["_start_dt"]) &
                (df_in["_end_dt"] == row["_end_dt"])
            )

            if col_event:
                mask &= (df_in[col_event] == row[col_event])

            matched = df_in[mask]
            if len(matched) <= 1:
                continue

            # check duplicated-market exemption
            chan_val = str(row[col_channel_id] if col_channel_id else row[col_channel]).lower().strip()
            markets = {str(m).lower().strip() for m in matched[col_market].fillna("")}

            if len(markets) > 1 and chan_val in {str(x).lower().strip() for x in duplicated_channels}:
                continue

            duplicate_ok.loc[matched.index] = False
            duplicate_remark.loc[matched.index] = "Duplicate row found (same market+channel+start+end[+event])"

    except Exception:
        logging.exception("Duplicate detection failed")
        duplicate_remark = duplicate_remark.where(~duplicate_remark.eq("OK"), "Duplicate detection error")

    # return combined results
    res_df = pd.DataFrame({
        "Overlap_OK": overlap_ok,
        "Overlap_Remark": overlap_remark,
        "Duplicate_OK": duplicate_ok,
        "Duplicate_Remark": duplicate_remark,
        "Daybreak_OK": daybreak_ok,
        "Daybreak_Remark": daybreak_remark
    }, index=df_in.index)

    df_out = df.join(res_df, how='left')

    for col in ["Overlap_OK", "Overlap_Remark", "Duplicate_OK", "Duplicate_Remark", "Daybreak_OK", "Daybreak_Remark"]:
        if col not in df_out.columns:
            df_out[col] = pd.NA

    return df_out
    
# ----------------------------12 Country Channel IDs Check-------------------------------
def country_channel_id_check(df, bsr_cols):
    
    df["Market_Channel_ID_OK"] = True
    df["Market_Channel_ID_Remark"] = "OK"

    ch_col = _find_column(df, bsr_cols['tv_channel'])
    ch_id_col = _find_column(df, bsr_cols['channel_id'])
    mkt_col = _find_column(df, bsr_cols['market'])
    mkt_id_col = _find_column(df, bsr_cols['market_id'])
    
    if not all([ch_col, ch_id_col, mkt_col, mkt_id_col]):
        logging.warning("ID Check: Missing one or more ID columns. Skipping.")
        df["Market_Channel_ID_OK"] = False
        df["Market_Channel_ID_Remark"] = "Check skipped: ID columns not found"
        return df

    def norm(x):
        return str(x).strip() if pd.notna(x) else ""

    channel_id_map = {}
    market_id_map = {}
    
    # Build maps first
    for idx, row in df.iterrows():
        channel = norm(row.get(ch_col))
        channel_id = norm(row.get(ch_id_col))
        market = norm(row.get(mkt_col))
        market_id = norm(row.get(mkt_id_col))

        if channel and channel_id and channel not in channel_id_map:
            channel_id_map[channel] = channel_id
        if market and market_id and market not in market_id_map:
            market_id_map[market] = market_id
            
    # Check for inconsistencies
    for idx, row in df.iterrows():
        channel = norm(row.get(ch_col))
        channel_id = norm(row.get(ch_id_col))
        market = norm(row.get(mkt_col))
        market_id = norm(row.get(mkt_id_col))
        
        remarks = []
        ok = True

        if channel and channel_id_map.get(channel) != channel_id:
            remarks.append(f"Channel '{channel}' has multiple IDs")
            ok = False
        if market and market_id_map.get(market) != market_id:
            remarks.append(f"Market '{market}' has multiple IDs")
            ok = False
            
        df.at[idx, "Market_Channel_ID_OK"] = ok
        df.at[idx, "Market_Channel_ID_Remark"] = "; ".join(remarks) if remarks else "OK"

    return df

# --------------------------13 Client Lstv OTT Check---------------------------------
def client_lstv_ott_check(df, bsr_cols, rules):
    
    df["Client_LSTV_OTT_OK"] = True
    df["Client_LSTV_OTT_Remark"] = "OK"
    
    ch_id_col = _find_column(df, bsr_cols['channel_id'])
    mkt_id_col = _find_column(df, bsr_cols['market_id'])
    pay_col = _find_column(df, bsr_cols['pay_tv'])
    keywords = rules.get('keywords', ['client', 'lstv', 'ott'])
    
    if not all([ch_id_col, mkt_id_col, pay_col]):
        logging.warning("Client Check: Missing ID or PayTV columns. Skipping.")
        df["Client_LSTV_OTT_OK"] = False
        df["Client_LSTV_OTT_Remark"] = "Check skipped: columns not found"
        return df
        
    def norm(x):
        return str(x).strip().lower() if pd.notna(x) else ""

    channel_to_market = {}
    market_to_channel = {}

    for idx, row in df.iterrows():
        ch_id = norm(row.get(ch_id_col))
        mk_id = norm(row.get(mkt_id_col))
        remarks = []
        ok = True

        if ch_id:
            if ch_id in channel_to_market and channel_to_market[ch_id] != mk_id:
                remarks.append(f"Channel ID {ch_id} linked to multiple Market IDs")
                ok = False
            else:
                channel_to_market[ch_id] = mk_id

        if mk_id:
            if mk_id in market_to_channel and market_to_channel[mk_id] != ch_id:
                remarks.append(f"Market ID {mk_id} linked to multiple Channel IDs")
                ok = False
            else:
                market_to_channel[mk_id] = ch_id

        val = norm(row.get(pay_col, ""))
        if not any(k in val for k in keywords):
            remarks.append(f"Missing Client/LSTV/OTT source: {row.get(pay_col, '')}")
            ok = False

        df.at[idx, "Client_LSTV_OTT_OK"] = ok
        df.at[idx, "Client_LSTV_OTT_Remark"] = "; ".join(remarks) if remarks else "OK"

    return df
# -----------------------------------------------------------
def color_excel(output_path, df):
    
    wb = load_workbook(output_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_map = {name: idx+1 for idx, name in enumerate(headers)}

    qc_columns = [col for col in df.columns if col.endswith("_OK")]

    for col_name in qc_columns:
        if col_name in col_map:
            col_idx = col_map[col_name]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                val_str = str(cell.value).lower()
                
                if val_str == "true":
                    cell.fill = GREEN_FILL
                elif val_str == "false":
                    cell.fill = RED_FILL
                # Cells with "Not Applicable" (pd.NA -> "nan") or "" remain uncolored

    wb.save(output_path)
# -----------------------------------------------------------
def generate_summary_sheet(output_path, df, file_rules):
    
    summary_sheet_name = file_rules.get('summary_sheet_name', 'Summary')
    
    wb = load_workbook(output_path)
    if summary_sheet_name in wb.sheetnames: 
        del wb[summary_sheet_name]
    ws = wb.create_sheet(summary_sheet_name)

    qc_columns = [col for col in df.columns if "_OK" in col]
    summary_data = []
    
    for col in qc_columns:
        total = len(df)
        
        # Convert to string and check, to handle bools and strings
        col_str = df[col].astype(str).str.lower()
        passed = (col_str == "true").sum()
        failed = (col_str == "false").sum()
        not_applicable = total - passed - failed
        
        summary_data.append([col, total, passed, failed, not_applicable])

    summary_df = pd.DataFrame(summary_data, columns=["Check", "Total", "Passed", "Failed", "N/A"])
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)
    wb.save(output_path)